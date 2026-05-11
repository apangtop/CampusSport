from rest_framework import viewsets, status, permissions
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
import io
from .models import Student, Registration, TeamRegistration
from .serializers import (
    StudentSerializer, RegistrationSerializer,
    RegistrationDetailSerializer, TeamRegistrationSerializer,
    BulkRegistrationSerializer
)
from events.models import Event, SportsMeet


class IsAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'admin'


class IsAdminOrTeacher(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ['admin', 'teacher']


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == 'teacher':
            qs = qs.filter(class_name=user.class_name)
        class_name = self.request.query_params.get('class_name')
        if class_name:
            if class_name.endswith('级') and not class_name.endswith('班'):
                qs = qs.filter(class_name__startswith=class_name)
            else:
                qs = qs.filter(class_name=class_name)
        name = self.request.query_params.get('name')
        if name:
            qs = qs.filter(name__icontains=name)
        gender = self.request.query_params.get('gender')
        if gender:
            qs = qs.filter(gender=gender)
        grade = self.request.query_params.get('grade')
        if grade:
            qs = qs.filter(grade=grade)
        return qs

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminOrTeacher()]
        return [permissions.IsAuthenticated()]

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'teacher':
            class_name = serializer.validated_data.get('class_name', '')
            if class_name != user.class_name:
                raise PermissionDenied('只能添加本班学生')
        serializer.save()

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrTeacher])
    def import_excel(self, request):
        """从 Excel 批量导入学生"""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        default_class = request.data.get('class_name') or request.user.class_name or ''
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': '文件格式错误，请上传xlsx文件'}, status=status.HTTP_400_BAD_REQUEST)

        def cell(row, idx, default=''):
            """安全读取单元格，None → 空字符串"""
            v = row[idx] if len(row) > idx else None
            return str(v).strip() if v is not None else default

        created, updated, errors = [], [], []
        import re
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            name = cell(row, 0)
            if not name:
                errors.append(f'第{row_idx}行：姓名为空')
                continue
            gender_raw = cell(row, 1, '男')
            student_id = cell(row, 2)
            grade = cell(row, 3)
            row_class = cell(row, 4)
            gender = 'male' if gender_raw in ['男', 'male', 'M', 'm'] else 'female'

            # 如果 Excel 行中没指定班级，使用请求中的默认班级
            student_class = row_class or default_class
            if not student_class:
                errors.append(f'第{row_idx}行：未指定班级')
                continue

            # 自动从班级名提取年级（如 2028级1班 → 2028级）
            if not grade and student_class:
                match = re.match(r'(\d{4}级)', student_class)
                if match:
                    grade = match.group(1)

            # 优先用学号+班级匹配，否则用姓名+班级
            if student_id:
                obj, is_created = Student.objects.update_or_create(
                    student_id=student_id, class_name=student_class,
                    defaults={'name': name, 'gender': gender, 'grade': grade}
                )
            else:
                obj, is_created = Student.objects.update_or_create(
                    name=name, class_name=student_class,
                    defaults={'gender': gender, 'student_id': student_id, 'grade': grade}
                )
            (created if is_created else updated).append(name)

        return Response({
            'created': len(created),
            'updated': len(updated),
            'errors': errors,
            'detail': f'导入完成：新增{len(created)}人，更新{len(updated)}人'
        })

    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrTeacher])
    def export_template(self, request):
        """下载学生导入模板"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '学生名单'
        headers = ['姓名', '性别(男/女)', '学号', '年级', '班级']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        # 示例数据
        ws.append(['张三', '男', '2024001', '2028级', '2028级1班'])
        ws.append(['李四', '女', '2024002', '2028级', '2028级1班'])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_template.xlsx"'
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrTeacher])
    def bulk(self, request):
        """批量创建学生"""
        students_data = request.data if isinstance(request.data, list) else request.data.get('students', [])
        if not students_data:
            return Response({'detail': '请提供学生列表'}, status=status.HTTP_400_BAD_REQUEST)

        class_name = request.data.get('class_name') if isinstance(request.data, dict) else ''
        created, errors = [], []
        for item in students_data:
            name = item.get('name', '').strip()
            if not name:
                errors.append('姓名为空')
                continue
            item_class = item.get('class_name', class_name) or (request.user.class_name if request.user.role == 'teacher' else '')
            if not item_class:
                errors.append(f'{name}: 缺少班级信息')
                continue
            gender_raw = item.get('gender', '男')
            gender = 'male' if str(gender_raw) in ['男', 'male', 'M', 'm'] else 'female'
            obj, is_new = Student.objects.get_or_create(
                name=name, class_name=item_class,
                defaults={
                    'gender': gender,
                    'student_id': item.get('student_id', ''),
                    'grade': item.get('grade', ''),
                }
            )
            if is_new:
                created.append(name)
        return Response({'created': len(created), 'errors': errors, 'detail': f'新增 {len(created)} 名学生'})


class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = Registration.objects.select_related('student', 'event', 'submitted_by', 'schedule').all()
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAdmin()]
        return [permissions.IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return RegistrationDetailSerializer
        return RegistrationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == 'teacher':
            qs = qs.filter(student__class_name=user.class_name)
        elif user.role == 'referee':
            qs = qs.filter(event__referee=user)

        event_id = self.request.query_params.get('event')
        if event_id:
            qs = qs.filter(event_id=event_id)
        class_name = self.request.query_params.get('class_name')
        if class_name:
            if class_name.endswith('级') and not class_name.endswith('班'):
                qs = qs.filter(student__class_name__startswith=class_name)
            else:
                qs = qs.filter(student__class_name=class_name)
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        meet_id = self.request.query_params.get('sports_meet')
        if meet_id:
            qs = qs.filter(event__sports_meet_id=meet_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(submitted_by=self.request.user)

    def perform_destroy(self, instance):
        if instance.event.sports_meet.status in ('ongoing', 'finished'):
            raise PermissionDenied('进行中或已结束的运动会不能删除报名')
        instance.delete()

    def _check_teacher_permission(self, registration):
        user = self.request.user
        if user.role == 'teacher' and registration.student.class_name != user.class_name:
            return False
        return True

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self._check_teacher_permission(instance):
            return Response({'detail': '只能操作本班学生的报名'}, status=status.HTTP_403_FORBIDDEN)
        if instance.event.sports_meet.status not in ['registration'] and request.user.role != 'admin':
            return Response({'detail': '报名截止后不可修改'}, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrTeacher])
    def bulk_register(self, request):
        """批量报名"""
        serializer = BulkRegistrationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        event_id = serializer.validated_data['event_id']
        student_ids = serializer.validated_data['student_ids']

        try:
            event = Event.objects.select_related('sports_meet').get(id=event_id)
        except Event.DoesNotExist:
            return Response({'detail': '项目不存在'}, status=status.HTTP_404_NOT_FOUND)

        meet = event.sports_meet

        # 检查运动会状态和报名截止时间
        if meet.status != 'registration' and request.user.role != 'admin':
            return Response({'detail': '当前不在报名阶段'}, status=status.HTTP_400_BAD_REQUEST)
        if meet.registration_deadline and timezone.now() > meet.registration_deadline and request.user.role != 'admin':
            return Response({'detail': '报名已截止'}, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        if user.role == 'teacher':
            students = Student.objects.filter(id__in=student_ids, class_name=user.class_name)
            if students.count() != len(student_ids):
                return Response({'detail': '只能为本班学生报名'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        errors = []
        with transaction.atomic():
            # 在事务内检查每班报名上限，防止竞态条件
            if user.role == 'teacher':
                current_count = Registration.objects.select_for_update().filter(
                    event=event,
                    student__class_name=user.class_name,
                    status__in=['submitted', 'approved']
                ).count()
                if current_count + len(student_ids) > event.max_per_class:
                    return Response(
                        {'detail': f'超出每班报名上限({event.max_per_class}人)'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)

                    # 检查性别匹配
                    if event.gender != 'mixed' and student.gender != event.gender:
                        errors.append(f'{student.name} 性别不符（项目要求{event.get_gender_display()}）')
                        continue

                    # 检查年级匹配
                    if event.grade and student.grade != event.grade:
                        errors.append(f'{student.name} 年级不符（项目要求{event.grade}，学生为{student.grade}）')
                        continue

                    # 检查每人报名项目数：以 event.max_per_person 为准，meet 上限兜底
                    per_event_cap = event.max_per_person or meet.max_events_per_person
                    person_count = Registration.objects.filter(
                        student=student,
                        event__sports_meet=meet,
                        status__in=['submitted', 'approved']
                    ).count()
                    if person_count >= per_event_cap:
                        errors.append(f'{student.name} 已达到报名项目上限（{per_event_cap}个）')
                        continue
                    reg, created_flag = Registration.objects.get_or_create(
                        event=event, student=student,
                        defaults={'submitted_by': user, 'status': 'submitted'}
                    )
                    if created_flag:
                        created.append(student.name)
                    elif reg.status == 'cancelled':
                        reg.status = 'submitted'
                        reg.submitted_by = user
                        reg.save()
                        created.append(student.name)
                except Student.DoesNotExist:
                    errors.append(f'学生ID {student_id} 不存在')

        return Response({'created': created, 'errors': errors})

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def approve(self, request, pk=None):
        reg = self.get_object()
        reg.status = 'approved'
        reg.save()
        return Response({'detail': '已审核通过'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def reject(self, request, pk=None):
        reg = self.get_object()
        reg.status = 'rejected'
        reg.save()
        return Response({'detail': '已拒绝'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrTeacher])
    def cancel(self, request, pk=None):
        reg = self.get_object()
        user = request.user
        if user.role == 'teacher' and reg.student.class_name != user.class_name:
            return Response({'detail': '只能取消本班报名'}, status=status.HTTP_403_FORBIDDEN)
        if reg.event.sports_meet.status not in ('registration',) and user.role != 'admin':
            return Response({'detail': '报名已截止，无法取消'}, status=status.HTTP_400_BAD_REQUEST)
        reg.status = 'cancelled'
        reg.save()
        return Response({'detail': '已取消报名'})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def approve_all(self, request):
        """一键审核所有提交的报名"""
        event_id = request.data.get('event_id')
        qs = Registration.objects.filter(status='submitted')
        if event_id:
            qs = qs.filter(event_id=event_id)
        count = qs.update(status='approved')
        return Response({'detail': f'已审核 {count} 条报名'})


class TeamRegistrationViewSet(viewsets.ModelViewSet):
    queryset = TeamRegistration.objects.prefetch_related('members').all()
    serializer_class = TeamRegistrationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update']:
            return [IsAdminOrTeacher()]
        if self.action == 'destroy':
            return [IsAdmin()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == 'teacher':
            qs = qs.filter(class_name=user.class_name)
        elif user.role == 'referee':
            qs = qs.filter(event__referee=user)
        event_id = self.request.query_params.get('event')
        if event_id:
            qs = qs.filter(event_id=event_id)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        class_name = serializer.validated_data.get('class_name', '')
        if user.role == 'teacher' and class_name != user.class_name:
            raise PermissionDenied('只能为本班报名团体项目')
        serializer.save(submitted_by=user)

    def perform_destroy(self, instance):
        if instance.event.sports_meet.status in ('ongoing', 'finished'):
            raise PermissionDenied('进行中或已结束的运动会不能删除团体报名')
        instance.delete()
