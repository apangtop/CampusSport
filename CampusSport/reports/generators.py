import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from django.conf import settings
from events.models import SportsMeet, Event
from registration.models import Registration, TeamRegistration


def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_order_book_excel(sports_meet_id):
    """生成秩序册 Excel"""
    try:
        meet = SportsMeet.objects.get(id=sports_meet_id)
    except SportsMeet.DoesNotExist:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    # ===== 封面页 =====
    cover = wb.create_sheet('封面')
    cover.column_dimensions['A'].width = 60
    cover['A1'] = f'第{meet.session}届 {meet.name}'
    cover['A1'].font = Font(name='黑体', size=22, bold=True)
    cover['A1'].alignment = Alignment(horizontal='center', vertical='center')
    cover.row_dimensions[1].height = 60

    if meet.school:
        cover['A2'] = meet.school
        cover['A2'].font = Font(size=14)
        cover['A2'].alignment = Alignment(horizontal='center')

    if meet.start_date:
        cover['A3'] = f'举办日期：{meet.start_date} 至 {meet.end_date or meet.start_date}'
        cover['A3'].alignment = Alignment(horizontal='center')

    cover['A4'] = '秩序册'
    cover['A4'].font = Font(size=16, bold=True)
    cover['A4'].alignment = Alignment(horizontal='center')

    # ===== 赛程总览页 =====
    overview = wb.create_sheet('赛程总览')
    headers = ['项目名称', '项目类型', '性别', '阶段', '组次', '比赛时间', '场地', '负责裁判']
    for col, h in enumerate(headers, 1):
        cell = overview.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

    col_widths = [20, 12, 8, 12, 8, 20, 15, 12]
    for i, w in enumerate(col_widths, 1):
        overview.column_dimensions[get_column_letter(i)].width = w

    row = 2
    events = meet.events.prefetch_related('schedules').select_related('referee').order_by('event_type', 'name')
    for event in events:
        schedules = event.schedules.all()
        if schedules.exists():
            for sch in schedules:
                overview.cell(row=row, column=1, value=event.name).border = thin_border()
                overview.cell(row=row, column=2, value=event.get_event_type_display()).border = thin_border()
                overview.cell(row=row, column=3, value=event.get_gender_display()).border = thin_border()
                overview.cell(row=row, column=4, value=sch.get_stage_display()).border = thin_border()
                overview.cell(row=row, column=5, value=sch.group_number).border = thin_border()
                time_val = str(sch.scheduled_time) if sch.scheduled_time else ''
                overview.cell(row=row, column=6, value=time_val).border = thin_border()
                overview.cell(row=row, column=7, value=sch.venue).border = thin_border()
                referee_name = event.referee.real_name if event.referee else ''
                overview.cell(row=row, column=8, value=referee_name).border = thin_border()
                for col in range(1, 9):
                    overview.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                row += 1
        else:
            overview.cell(row=row, column=1, value=event.name).border = thin_border()
            overview.cell(row=row, column=2, value=event.get_event_type_display()).border = thin_border()
            overview.cell(row=row, column=3, value=event.get_gender_display()).border = thin_border()
            referee_name = event.referee.real_name if event.referee else ''
            overview.cell(row=row, column=8, value=referee_name).border = thin_border()
            row += 1

    # ===== 各项目参赛名单页 =====
    for event in events:
        is_team = event.event_type in ['team_confrontation', 'relay']
        sheet_name = event.name[:25]
        ws = wb.create_sheet(sheet_name)

        # 标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'【{event.name}】参赛名单'
        title_cell.font = Font(name='黑体', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # 项目信息
        ws['A2'] = f'项目类型：{event.get_event_type_display()}'
        ws['C2'] = f'参赛性别：{event.get_gender_display()}'
        ws['E2'] = f'负责裁判：{event.referee.real_name if event.referee else "未分配"}'
        for cell in [ws['A2'], ws['C2'], ws['E2']]:
            cell.font = Font(size=10)

        if is_team:
            headers = ['序号', '班级', '队员名单', '成绩', '名次', '备注']
        else:
            headers = ['道次', '姓名', '班级', '成绩', '名次', '备注']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 15

        row = 4
        if is_team:
            team_regs = TeamRegistration.objects.filter(
                event=event, status__in=['submitted', 'approved']
            ).prefetch_related('members').order_by('lane', 'class_name')
            for idx, team_reg in enumerate(team_regs, 1):
                members_str = '、'.join([m.name for m in team_reg.members.all()])
                ws.cell(row=row, column=1, value=team_reg.lane or idx).border = thin_border()
                ws.cell(row=row, column=2, value=team_reg.class_name).border = thin_border()
                ws.cell(row=row, column=3, value=members_str).border = thin_border()
                ws.cell(row=row, column=4, value='').border = thin_border()
                ws.cell(row=row, column=5, value='').border = thin_border()
                ws.cell(row=row, column=6, value='').border = thin_border()
                for col in range(1, 7):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                row += 1
        else:
            regs = Registration.objects.filter(
                event=event, status__in=['submitted', 'approved']
            ).select_related('student', 'schedule').order_by('lane', 'student__class_name')
            for reg in regs:
                ws.cell(row=row, column=1, value=reg.lane or '').border = thin_border()
                ws.cell(row=row, column=2, value=reg.student.name).border = thin_border()
                ws.cell(row=row, column=3, value=reg.student.class_name).border = thin_border()
                ws.cell(row=row, column=4, value='').border = thin_border()
                ws.cell(row=row, column=5, value='').border = thin_border()
                ws.cell(row=row, column=6, value='').border = thin_border()
                for col in range(1, 7):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                row += 1

        if row == 4:
            ws.cell(row=4, column=1, value='暂无参赛名单')

    # ===== 积分规则页 =====
    rules_ws = wb.create_sheet('积分规则')
    rules_ws['A1'] = '积分规则说明'
    rules_ws['A1'].font = Font(bold=True, size=13)
    rules_ws.merge_cells('A1:B1')
    rules_ws['A1'].alignment = Alignment(horizontal='center')

    rules_ws['A2'] = '名次'
    rules_ws['B2'] = '积分'
    rules_ws['A2'].font = Font(bold=True)
    rules_ws['B2'].font = Font(bold=True)

    default_rules = [(1, 7), (2, 5), (3, 4), (4, 3), (5, 2), (6, 1)]
    for i, (rank, pts) in enumerate(default_rules, 3):
        rules_ws.cell(row=i, column=1, value=f'第{rank}名').alignment = Alignment(horizontal='center')
        rules_ws.cell(row=i, column=2, value=pts).alignment = Alignment(horizontal='center')
    rules_ws.column_dimensions['A'].width = 15
    rules_ws.column_dimensions['B'].width = 10

    # 保存文件
    media_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(media_dir, exist_ok=True)
    filename = f'order_book_{sports_meet_id}.xlsx'
    filepath = os.path.join(media_dir, filename)
    wb.save(filepath)
    return f'reports/{filename}'


def generate_result_report_excel(sports_meet_id):
    """生成成绩报表 Excel"""
    try:
        meet = SportsMeet.objects.get(id=sports_meet_id)
    except SportsMeet.DoesNotExist:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    # ===== 班级积分榜 =====
    from scores.models import ClassPoints
    points_ws = wb.create_sheet('班级积分榜')
    headers = ['排名', '班级', '总积分', '金牌', '银牌', '铜牌']
    for col, h in enumerate(headers, 1):
        cell = points_ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='D97706')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    class_points = ClassPoints.objects.filter(sports_meet=meet).order_by('rank')
    for idx, cp in enumerate(class_points, 2):
        row_data = [cp.rank, cp.class_name, cp.total_points, cp.gold_medals, cp.silver_medals, cp.bronze_medals]
        for col, val in enumerate(row_data, 1):
            cell = points_ws.cell(row=idx, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
            if idx == 2:
                cell.fill = PatternFill(fill_type='solid', fgColor='FEF3C7')

    for i, w in enumerate([8, 15, 10, 8, 8, 8], 1):
        points_ws.column_dimensions[get_column_letter(i)].width = w

    # ===== 各项目成绩页 =====
    events = meet.events.prefetch_related('schedules').select_related('referee').order_by('event_type', 'name')
    for event in events:
        sheet_name = f'{event.name[:20]}_成绩'
        ws = wb.create_sheet(sheet_name)
        is_team = event.event_type in ['team_confrontation', 'relay']

        ws.merge_cells('A1:F1')
        ws['A1'] = f'{event.name} 最终成绩'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')

        if is_team:
            headers = ['名次', '班级', '成绩', '积分']
        else:
            headers = ['名次', '姓名', '班级', '成绩', '积分']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()

        from scores.models import Score, TeamScore
        row = 3
        if is_team:
            team_scores = TeamScore.objects.filter(
                team_registration__event=event, stage='final'
            ).select_related('team_registration').order_by('rank')
            for ts in team_scores:
                ws.cell(row=row, column=1, value=ts.rank).border = thin_border()
                ws.cell(row=row, column=2, value=ts.team_registration.class_name).border = thin_border()
                ws.cell(row=row, column=3, value=ts.result).border = thin_border()
                ws.cell(row=row, column=4, value=ts.points).border = thin_border()
                for col in range(1, 5):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                row += 1
        else:
            scores = Score.objects.filter(
                registration__event=event, stage='final'
            ).select_related('registration__student').order_by('rank')
            for s in scores:
                ws.cell(row=row, column=1, value=s.rank).border = thin_border()
                ws.cell(row=row, column=2, value=s.registration.student.name).border = thin_border()
                ws.cell(row=row, column=3, value=s.registration.student.class_name).border = thin_border()
                ws.cell(row=row, column=4, value=s.result).border = thin_border()
                ws.cell(row=row, column=5, value=s.points).border = thin_border()
                for col in range(1, 6):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                row += 1

        for i, w in enumerate([8, 15, 15, 12, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    media_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(media_dir, exist_ok=True)
    filename = f'result_report_{sports_meet_id}.xlsx'
    filepath = os.path.join(media_dir, filename)
    wb.save(filepath)
    return f'reports/{filename}'
